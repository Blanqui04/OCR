#!/usr/bin/env python3
"""
Data Export Utilities for OCR Viewer Application
Handles exporting OCR results to various formats
"""

import json
import csv
import logging
from typing import Dict, List, Optional, Any
from pathlib import Path
from datetime import datetime
import io

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
except ImportError:
    SimpleDocTemplate = None

logger = logging.getLogger(__name__)

class DataExporter:
    """Handles data export functionality"""
    
    def __init__(self):
        """Initialize the data exporter"""
        self.styles = None
        if SimpleDocTemplate:
            self.styles = getSampleStyleSheet()
            
    def export_text(self, ocr_results: Dict, file_path: str) -> bool:
        """
        Export extracted text to a text file
        
        Args:
            ocr_results: OCR results dictionary
            file_path: Output file path
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                # Write header
                f.write("OCR Extracted Text\n")
                f.write("=" * 50 + "\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                # Write statistics
                stats = ocr_results.get('statistics', {})
                f.write(f"Total pages: {stats.get('page_count', 0)}\n")
                f.write(f"Total text blocks: {stats.get('total_blocks', 0)}\n")
                f.write(f"Average confidence: {stats.get('avg_confidence', 0):.2%}\n")
                
                languages = stats.get('languages', [])
                if languages:
                    f.write(f"Detected languages: {', '.join(languages)}\n")
                    
                f.write("\n" + "=" * 50 + "\n\n")
                
                # Write full text
                full_text = ocr_results.get('text', '')
                if full_text:
                    f.write("FULL TEXT:\n\n")
                    f.write(full_text)
                else:
                    # If no full text, combine text blocks
                    f.write("TEXT BY BLOCKS:\n\n")
                    blocks = ocr_results.get('blocks', [])
                    for i, block in enumerate(blocks, 1):
                        text = block.get('text', '').strip()
                        if text:
                            confidence = block.get('confidence', 0)
                            f.write(f"Block {i} (Confidence: {confidence:.2%}):\n")
                            f.write(f"{text}\n\n")
                            
            logger.info(f"Text exported successfully to: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting text: {str(e)}")
            return False
            
    def export_json(self, ocr_results: Dict, file_path: str) -> bool:
        """
        Export OCR results to JSON format
        
        Args:
            ocr_results: OCR results dictionary
            file_path: Output file path
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Prepare export data
            export_data = {
                'metadata': {
                    'export_time': datetime.now().isoformat(),
                    'version': '2.0',
                    'description': 'OCR results from OCR Viewer Application'
                },
                'statistics': ocr_results.get('statistics', {}),
                'text': ocr_results.get('text', ''),
                'blocks': ocr_results.get('blocks', []),
                'pages': ocr_results.get('pages', [])
            }
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
                
            logger.info(f"JSON exported successfully to: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting JSON: {str(e)}")
            return False
            
    def export_csv(self, ocr_results: Dict, file_path: str) -> bool:
        """
        Export text blocks to CSV format
        
        Args:
            ocr_results: OCR results dictionary
            file_path: Output file path
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            blocks = ocr_results.get('blocks', [])
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write header
                writer.writerow([
                    'Block ID',
                    'Page',
                    'Text',
                    'Confidence',
                    'X',
                    'Y',
                    'Width',
                    'Height',
                    'Character Count',
                    'Word Count'
                ])
                
                # Write data rows
                for block in blocks:
                    text = block.get('text', '').strip()
                    writer.writerow([
                        block.get('block_id', ''),
                        block.get('page', 1),
                        text,
                        f"{block.get('confidence', 0):.4f}",
                        f"{block.get('x', 0):.6f}",
                        f"{block.get('y', 0):.6f}",
                        f"{block.get('width', 0):.6f}",
                        f"{block.get('height', 0):.6f}",
                        len(text),
                        len(text.split()) if text else 0
                    ])
                    
            logger.info(f"CSV exported successfully to: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting CSV: {str(e)}")
            return False
            
    def export_pdf_report(self, ocr_results: Dict, file_path: str) -> bool:
        """
        Export a comprehensive PDF report
        
        Args:
            ocr_results: OCR results dictionary
            file_path: Output file path
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            if not SimpleDocTemplate:
                logger.error("ReportLab not installed. Cannot export PDF reports.")
                return False
                
            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Build content
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph("OCR Analysis Report", title_style))
            story.append(Spacer(1, 12))
            
            # Metadata
            story.append(Paragraph("Report Information", self.styles['Heading2']))
            metadata_table = [
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Application:', 'OCR Viewer Professional v2.0']
            ]
            
            stats = ocr_results.get('statistics', {})
            metadata_table.extend([
                ['Total Pages:', str(stats.get('page_count', 0))],
                ['Total Text Blocks:', str(stats.get('total_blocks', 0))],
                ['Average Confidence:', f"{stats.get('avg_confidence', 0):.2%}"]
            ])
            
            languages = stats.get('languages', [])
            if languages:
                metadata_table.append(['Detected Languages:', ', '.join(languages)])
                
            t = Table(metadata_table, colWidths=[2*inch, 4*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(t)
            story.append(Spacer(1, 12))
            
            # Confidence Distribution
            story.append(Paragraph("Confidence Analysis", self.styles['Heading2']))
            
            blocks = ocr_results.get('blocks', [])
            if blocks:
                high_conf = sum(1 for b in blocks if b.get('confidence', 0) >= 0.9)
                med_conf = sum(1 for b in blocks if 0.7 <= b.get('confidence', 0) < 0.9)
                low_conf = sum(1 for b in blocks if b.get('confidence', 0) < 0.7)
                
                conf_table = [
                    ['Confidence Level', 'Count', 'Percentage'],
                    ['High (≥90%)', str(high_conf), f"{high_conf/len(blocks)*100:.1f}%"],
                    ['Medium (70-90%)', str(med_conf), f"{med_conf/len(blocks)*100:.1f}%"],
                    ['Low (<70%)', str(low_conf), f"{low_conf/len(blocks)*100:.1f}%"]
                ]
                
                conf_table_obj = Table(conf_table, colWidths=[2*inch, 1*inch, 1*inch])
                conf_table_obj.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(conf_table_obj)
                story.append(Spacer(1, 12))
                
            # Sample text blocks
            story.append(Paragraph("Sample Text Blocks", self.styles['Heading2']))
            
            sample_blocks = blocks[:10]  # First 10 blocks
            for i, block in enumerate(sample_blocks, 1):
                text = block.get('text', '').strip()
                if text:
                    confidence = block.get('confidence', 0)
                    block_title = f"Block {i} (Page {block.get('page', 1)}, Confidence: {confidence:.2%})"
                    story.append(Paragraph(block_title, self.styles['Heading3']))
                    
                    # Truncate long text
                    display_text = text[:200] + "..." if len(text) > 200 else text
                    story.append(Paragraph(display_text, self.styles['Normal']))
                    story.append(Spacer(1, 6))
                    
            # Build PDF
            doc.build(story)
            
            logger.info(f"PDF report exported successfully to: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting PDF report: {str(e)}")
            return False
            
    def get_export_statistics(self, ocr_results: Dict) -> Dict:
        """
        Generate comprehensive statistics for export
        
        Args:
            ocr_results: OCR results dictionary
            
        Returns:
            Dict: Detailed statistics
        """
        try:
            blocks = ocr_results.get('blocks', [])
            text = ocr_results.get('text', '')
            
            if not blocks:
                return {}
                
            # Basic statistics
            total_blocks = len(blocks)
            total_characters = sum(len(block.get('text', '')) for block in blocks)
            total_words = sum(len(block.get('text', '').split()) for block in blocks)
            
            # Confidence statistics
            confidences = [block.get('confidence', 0) for block in blocks]
            avg_confidence = sum(confidences) / len(confidences) if confidences else 0
            min_confidence = min(confidences) if confidences else 0
            max_confidence = max(confidences) if confidences else 0
            
            # Confidence distribution
            high_conf_blocks = sum(1 for c in confidences if c >= 0.9)
            med_conf_blocks = sum(1 for c in confidences if 0.7 <= c < 0.9)
            low_conf_blocks = sum(1 for c in confidences if c < 0.7)
            
            # Page statistics
            pages = set(block.get('page', 1) for block in blocks)
            page_count = len(pages)
            
            # Size statistics
            widths = [block.get('width', 0) for block in blocks]
            heights = [block.get('height', 0) for block in blocks]
            
            avg_width = sum(widths) / len(widths) if widths else 0
            avg_height = sum(heights) / len(heights) if heights else 0
            
            statistics = {
                'total_blocks': total_blocks,
                'total_characters': total_characters,
                'total_words': total_words,
                'page_count': page_count,
                'confidence': {
                    'average': avg_confidence,
                    'minimum': min_confidence,
                    'maximum': max_confidence,
                    'high_confidence_blocks': high_conf_blocks,
                    'medium_confidence_blocks': med_conf_blocks,
                    'low_confidence_blocks': low_conf_blocks
                },
                'dimensions': {
                    'average_width': avg_width,
                    'average_height': avg_height
                }
            }
            
            return statistics
            
        except Exception as e:
            logger.error(f"Error generating statistics: {str(e)}")
            return {}
    
    def export_structured_csv(self, relations: List[Dict], file_path: str) -> bool:
        """
        Export structured element-value relations to CSV format
        
        Args:
            relations: List of element-value relation dictionaries
            file_path: Output file path
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write header
                writer.writerow(['Element', 'Value', 'Confidence', 'Source', 'Created'])
                
                # Write data rows
                for relation in relations:
                    confidence = relation.get('confidence', 0.0)
                    confidence_str = f"{confidence:.4f}" if confidence > 0 else "N/A"
                    
                    writer.writerow([
                        relation.get('element', ''),
                        relation.get('value', ''),
                        confidence_str,
                        relation.get('source', 'Manual'),
                        relation.get('created', datetime.now().isoformat())
                    ])
            
            logger.info(f"Structured CSV exported successfully to: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting structured CSV: {str(e)}")
            return False
    
    def export_structured_excel(self, relations: List[Dict], file_path: str, metadata: Dict = None) -> bool:
        """
        Export structured element-value relations to Excel format
        
        Args:
            relations: List of element-value relation dictionaries
            file_path: Output file path
            metadata: Optional metadata dictionary
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Try to import openpyxl for Excel export
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                logger.error("openpyxl not available for Excel export")
                return False
            
            wb = Workbook()
            
            # Main relations sheet
            ws_relations = wb.active
            ws_relations.title = "Element_Value_Relations"
            
            # Headers
            headers = ['Element', 'Value', 'Confidence', 'Source', 'Created']
            for col, header in enumerate(headers, 1):
                cell = ws_relations.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row, relation in enumerate(relations, 2):
                confidence = relation.get('confidence', 0.0)
                confidence_str = f"{confidence:.4f}" if confidence > 0 else "N/A"
                
                ws_relations.cell(row=row, column=1, value=relation.get('element', ''))
                ws_relations.cell(row=row, column=2, value=relation.get('value', ''))
                ws_relations.cell(row=row, column=3, value=confidence_str)
                ws_relations.cell(row=row, column=4, value=relation.get('source', 'Manual'))
                ws_relations.cell(row=row, column=5, value=relation.get('created', datetime.now().isoformat()))
                
                # Color coding based on confidence
                if confidence >= 0.8:
                    fill_color = "E8F5E8"  # Light green
                elif confidence >= 0.6:
                    fill_color = "FFF4E6"  # Light orange
                else:
                    fill_color = "FFE6E6"  # Light red
                
                for col in range(1, 6):
                    ws_relations.cell(row=row, column=col).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
            
            # Auto-adjust column widths
            for col in range(1, 6):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in ws_relations[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_relations.column_dimensions[column_letter].width = adjusted_width
            
            # Metadata sheet
            if metadata:
                ws_metadata = wb.create_sheet("Metadata")
                
                # Metadata headers
                ws_metadata.cell(row=1, column=1, value="Property").font = Font(bold=True)
                ws_metadata.cell(row=1, column=2, value="Value").font = Font(bold=True)
                
                # Metadata rows
                row = 2
                for key, value in metadata.items():
                    ws_metadata.cell(row=row, column=1, value=str(key))
                    ws_metadata.cell(row=row, column=2, value=str(value))
                    row += 1
                
                # Auto-adjust metadata columns
                for col in [1, 2]:
                    column_letter = get_column_letter(col)
                    ws_metadata.column_dimensions[column_letter].width = 20
            
            # Summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Structured Data Export Summary").font = Font(size=16, bold=True)
            ws_summary.cell(row=3, column=1, value="Total Relations:").font = Font(bold=True)
            ws_summary.cell(row=3, column=2, value=len(relations))
            
            # Confidence distribution
            high_conf = sum(1 for r in relations if r.get('confidence', 0) >= 0.8)
            med_conf = sum(1 for r in relations if 0.6 <= r.get('confidence', 0) < 0.8)
            low_conf = sum(1 for r in relations if r.get('confidence', 0) < 0.6)
            
            ws_summary.cell(row=5, column=1, value="Confidence Distribution:").font = Font(bold=True)
            ws_summary.cell(row=6, column=1, value="High (≥80%):")
            ws_summary.cell(row=6, column=2, value=high_conf)
            ws_summary.cell(row=7, column=1, value="Medium (60-79%):")
            ws_summary.cell(row=7, column=2, value=med_conf)
            ws_summary.cell(row=8, column=1, value="Low (<60%):")
            ws_summary.cell(row=8, column=2, value=low_conf)
            
            ws_summary.cell(row=10, column=1, value="Export Date:").font = Font(bold=True)
            ws_summary.cell(row=10, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            # Save the workbook
            wb.save(file_path)
            
            logger.info(f"Structured Excel exported successfully to: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting structured Excel: {str(e)}")
            return False
